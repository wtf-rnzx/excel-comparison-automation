"""Compare two CSV/Excel files by fullname/FULLNAME column.

Recommended structure for scalability:
	data/source/   -> source-of-truth files
	data/target/   -> target files to compare
	output/        -> generated reports

Usage:
	python automate.py --source source.xlsx --target candidate.xlsx
	python automate.py --source source.csv --target candidate.csv

Optional examples:
	python automate.py --source source.xlsx --target candidate.xlsx --source-column FULLNAME --target-column fullname
	python automate.py --source source.xlsx --target candidate.xlsx --sheet-source Employees --sheet-target Employees
	python automate.py --source source.csv --target candidate.csv --output comparison_result.xlsx
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Iterable

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
SOURCE_DATA_DIR = BASE_DIR / "data" / "source"
TARGET_DATA_DIR = BASE_DIR / "data" / "target"
OUTPUT_DIR = BASE_DIR / "output"
SUPPORTED_TABULAR_EXTENSIONS = {".csv", ".txt", ".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"}


DEFAULT_NAME_CANDIDATES = ("fullname", "full name", "FULLNAME", "FULL NAME")
DEFAULT_CAMPUS_CANDIDATES = ("campus", "CAMPUS")
DEFAULT_SOURCE_PROGRAM_CANDIDATES = ("program_name", "program name", "PROGRAM_NAME", "PROGRAM NAME")
DEFAULT_TARGET_PROGRAM_CANDIDATES = ("program", "PROGRAM", "program_name", "program name")
DEFAULT_SOURCE_COLLEGE_CANDIDATES = ("college", "COLLEGE")
DEFAULT_TARGET_UNDERGRAD_COLLEGE_CANDIDATES = (
	"undergraduate_college",
	"undergraduate college",
	"undergraduate_collge",
	"undergraduate collge",
	"UNDERGRADUATE_COLLEGE",
	"UNDERGRADUATE COLLEGE",
	"college",
)
DEFAULT_TARGET_GRADUATE_COLLEGE_CANDIDATES = (
	"graduate_college",
	"graduate college",
	"graduate_collge",
	"graduate collge",
	"GRADUATE_COLLEGE",
	"GRADUATE COLLEGE",
	"college",
)
DEFAULT_TARGET_COLLEGE_CANDIDATES = (
	*DEFAULT_TARGET_UNDERGRAD_COLLEGE_CANDIDATES,
	*DEFAULT_TARGET_GRADUATE_COLLEGE_CANDIDATES,
	"COLLEGE",
)

CAMPUS_EQUIVALENCY_MAP = {
	"MAIN": "MAIN",
	"PABLO BORBON": "MAIN",
	"MALVAR": "MALVAR",
	"JPLPC-MALVAR": "MALVAR",
	"NASUGBU": "NASUGBU",
	"ARASOF-NASUGBU": "NASUGBU",
}

PROGRAM_EQUIVALENCY_MAP = {
	"BS HOSPITALITY MANAGEMENT": "BACHELOR OF SCIENCE IN HOSPITALITY MANAGEMENT",
	"BACHELOR OF SCIENCE IN HOSPITALITY MANAGEMENT": "BACHELOR OF SCIENCE IN HOSPITALITY MANAGEMENT",
}


@dataclass
class CompareResult:
	source_total_rows: int
	target_total_rows: int
	source_unique_names: int
	target_unique_names: int
	found_count: int
	not_found_count: int
	source_duplicate_names: int
	target_duplicate_names: int
	source_duplicate_rows: pd.DataFrame
	target_duplicate_rows: pd.DataFrame
	data_validation_rows: pd.DataFrame
	source_centered_rows: pd.DataFrame
	target_centered_rows: pd.DataFrame
	found_names: list[str]
	not_found_names: list[str]


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description=(
			"Compare a target Excel file against a source-of-truth Excel file "
			"using fullname/FULLNAME values."
		)
	)
	parser.add_argument("--source", required=True, help="Path to source-of-truth Excel file")
	parser.add_argument("--target", required=True, help="Path to target Excel file to compare")
	parser.add_argument(
		"--sheet-source",
		default=0,
		help="Source sheet name or index (default: first sheet)",
	)
	parser.add_argument(
		"--sheet-target",
		default=0,
		help="Target sheet name or index (default: first sheet)",
	)
	parser.add_argument(
		"--source-column",
		default=None,
		help="Column name in source file (default: auto-detect fullname/FULLNAME)",
	)
	parser.add_argument(
		"--target-column",
		default=None,
		help="Column name in target file (default: auto-detect fullname/FULLNAME)",
	)
	parser.add_argument(
		"--output",
		default=None,
		help=(
			"Output report path (.xlsx). If omitted, a timestamped file is created "
			"in output/."
		),
	)
	return parser.parse_args()


def ensure_managed_directories() -> None:
	SOURCE_DATA_DIR.mkdir(parents=True, exist_ok=True)
	TARGET_DATA_DIR.mkdir(parents=True, exist_ok=True)
	OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def resolve_input_path(path_value: str, preferred_dir: Path) -> Path:
	"""Resolve input path.

	Priority:
	1) exact provided path
	2) preferred managed folder + file name (when input is a simple file name)
	3) same file stem in preferred managed folder (extension may differ)
	"""
	provided = Path(path_value)
	if provided.exists():
		return provided

	if not provided.is_absolute() and provided.parent == Path("."):
		candidate = preferred_dir / provided.name
		if candidate.exists():
			return candidate

		# If user provides a different extension (e.g., .xlsx while actual is .csv),
		# search by file stem inside the managed folder.
		for file_path in preferred_dir.glob("*"):
			if not file_path.is_file():
				continue
			if file_path.suffix.lower() not in SUPPORTED_TABULAR_EXTENSIONS:
				continue
			if file_path.stem.strip().lower() == provided.stem.strip().lower():
				return file_path

	return provided


def resolve_output_path(path_value: str | None) -> Path:
	if not path_value:
		return build_default_output_path()

	provided = Path(path_value)
	if not provided.is_absolute() and provided.parent == Path("."):
		return OUTPUT_DIR / provided.name
	return provided


def normalize_name(value: object) -> str | None:
	if value is None or pd.isna(value):
		return None
	text = str(value).strip()
	if not text:
		return None
	normalized = " ".join(text.split()).upper()
	return normalized


def normalize_campus(value: object) -> str | None:
	normalized = normalize_name(value)
	if normalized is None:
		return None
	return CAMPUS_EQUIVALENCY_MAP.get(normalized, normalized)


def normalize_program(value: object) -> str | None:
	normalized = normalize_name(value)
	if normalized is None:
		return None

	if normalized in PROGRAM_EQUIVALENCY_MAP:
		return PROGRAM_EQUIVALENCY_MAP[normalized]

	# Treat BS-prefixed programs as equivalent to "BACHELOR OF SCIENCE ..."
	if normalized == "BS":
		return "BACHELOR OF SCIENCE"

	bs_prefixed = re.match(r"^BS[\s\.-]+(.+)$", normalized)
	if bs_prefixed:
		remainder = bs_prefixed.group(1).strip()
		if remainder.startswith("IN "):
			return f"BACHELOR OF SCIENCE {remainder}"
		return f"BACHELOR OF SCIENCE IN {remainder}"

	return normalized


def resolve_column(df: pd.DataFrame, preferred: str | None) -> str:
	if preferred:
		for col in df.columns:
			if str(col).strip().lower() == preferred.strip().lower():
				return str(col)
		raise ValueError(f"Column '{preferred}' was not found. Available columns: {list(df.columns)}")

	normalized_lookup = {str(col).strip().lower(): str(col) for col in df.columns}
	for candidate in DEFAULT_NAME_CANDIDATES:
		key = candidate.strip().lower()
		if key in normalized_lookup:
			return normalized_lookup[key]

	raise ValueError(
		"Could not auto-detect fullname column. "
		f"Please provide --source-column/--target-column. Available columns: {list(df.columns)}"
	)


def resolve_column_with_candidates(
	df: pd.DataFrame,
	preferred: str | None,
	candidates: tuple[str, ...],
	error_hint: str,
) -> str:
	if preferred:
		for col in df.columns:
			if str(col).strip().lower() == preferred.strip().lower():
				return str(col)
		raise ValueError(f"Column '{preferred}' was not found. Available columns: {list(df.columns)}")

	normalized_lookup = {str(col).strip().lower(): str(col) for col in df.columns}
	for candidate in candidates:
		key = candidate.strip().lower()
		if key in normalized_lookup:
			return normalized_lookup[key]

	raise ValueError(f"{error_hint}. Available columns: {list(df.columns)}")


def resolve_columns_with_candidates(df: pd.DataFrame, candidates: tuple[str, ...]) -> list[str]:
	normalized_lookup = {str(col).strip().lower(): str(col) for col in df.columns}
	resolved: list[str] = []
	seen: set[str] = set()
	for candidate in candidates:
		key = candidate.strip().lower()
		if key in normalized_lookup:
			column_name = normalized_lookup[key]
			if column_name not in seen:
				resolved.append(column_name)
				seen.add(column_name)
	return resolved


def extract_normalized_names(df: pd.DataFrame, column: str) -> list[str]:
	names = [normalize_name(value) for value in df[column]]
	return [name for name in names if name is not None]


def count_duplicates(values: Iterable[str]) -> int:
	counter = Counter(values)
	return sum(1 for value_count in counter.values() if value_count > 1)


def get_duplicate_rows(df: pd.DataFrame, column: str) -> pd.DataFrame:
	rows = df.copy()
	rows.insert(0, "file_row_number", rows.index + 2)
	rows["normalized_fullname"] = rows[column].apply(normalize_name)

	duplicate_mask = rows["normalized_fullname"].notna() & rows.duplicated(
		subset=["normalized_fullname"], keep=False
	)
	duplicate_rows = rows.loc[duplicate_mask].copy()

	if duplicate_rows.empty:
		return pd.DataFrame(columns=["file_row_number", "normalized_fullname", "duplicate_count", *df.columns])

	duplicate_rows["duplicate_count"] = duplicate_rows.groupby("normalized_fullname")[
		"normalized_fullname"
	].transform("size")

	return duplicate_rows.sort_values(
		by=["normalized_fullname", "file_row_number"],
		ascending=[True, True],
	)


def get_normalized_set(values: pd.Series, normalizer=normalize_name) -> list[str]:
	normalized = [normalizer(value) for value in values]
	return sorted({item for item in normalized if item is not None})


def status_from_sets(left: list[str], right: list[str]) -> str:
	if not left or not right:
		return "MISSING"
	if set(left).intersection(set(right)):
		return "MATCH"
	return "MISMATCH"


def join_values(values: list[str]) -> str:
	return " | ".join(values) if values else ""


def build_data_validation_rows(
	source_df: pd.DataFrame,
	target_df: pd.DataFrame,
	source_name_col: str,
	target_name_col: str,
	source_campus_col: str,
	target_campus_col: str,
	source_program_col: str,
	target_program_col: str,
	source_college_col: str,
	target_college_cols: list[str],
	found_names: list[str],
) -> pd.DataFrame:
	source_enriched = source_df.copy()
	target_enriched = target_df.copy()
	source_enriched["_normalized_fullname"] = source_enriched[source_name_col].apply(normalize_name)
	target_enriched["_normalized_fullname"] = target_enriched[target_name_col].apply(normalize_name)

	rows: list[dict[str, object]] = []
	for matched_name in found_names:
		source_rows = source_enriched[source_enriched["_normalized_fullname"] == matched_name]
		target_rows = target_enriched[target_enriched["_normalized_fullname"] == matched_name]

		source_campus_values = get_normalized_set(source_rows[source_campus_col], normalizer=normalize_campus)
		target_campus_values = get_normalized_set(target_rows[target_campus_col], normalizer=normalize_campus)
		source_program_values = get_normalized_set(source_rows[source_program_col], normalizer=normalize_program)
		target_program_values = get_normalized_set(target_rows[target_program_col], normalizer=normalize_program)
		source_college_values = get_normalized_set(source_rows[source_college_col])
		target_college_set: set[str] = set()
		for target_college_col in target_college_cols:
			target_college_set.update(get_normalized_set(target_rows[target_college_col]))
		target_college_values = sorted(target_college_set)

		campus_status = status_from_sets(source_campus_values, target_campus_values)
		program_status = status_from_sets(source_program_values, target_program_values)
		college_status = status_from_sets(source_college_values, target_college_values)

		if campus_status == "MATCH" and program_status == "MATCH" and college_status == "MATCH":
			overall_status = "MATCH"
		elif campus_status == "MISSING" or program_status == "MISSING" or college_status == "MISSING":
			overall_status = "REVIEW"
		else:
			overall_status = "MISMATCH"

		rows.append(
			{
				"fullname": matched_name,
				"source_rows_count": len(source_rows),
				"target_rows_count": len(target_rows),
				"source_campus": join_values(source_campus_values),
				"target_campus": join_values(target_campus_values),
				"campus_validation": campus_status,
				"source_program_name": join_values(source_program_values),
				"target_program": join_values(target_program_values),
				"program_validation": program_status,
				"source_college": join_values(source_college_values),
				"target_college": join_values(target_college_values),
				"college_validation": college_status,
				"overall_validation": overall_status,
			}
		)

	return pd.DataFrame(rows)


def read_tabular_file(path: Path, sheet: str | int) -> pd.DataFrame:
	"""Read CSV or Excel file into a DataFrame.

	- CSV/TXT files are loaded with pandas.read_csv (sheet argument ignored).
	- Excel files are loaded with pandas.read_excel using the provided sheet.
	"""
	ext = path.suffix.lower()
	if ext in {".csv", ".txt"}:
		return pd.read_csv(path, dtype=str)
	if ext in {".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"}:
		try:
			if ext in {".xlsx", ".xlsm"}:
				return pd.read_excel(path, sheet_name=sheet, dtype=str, engine="openpyxl")
			return pd.read_excel(path, sheet_name=sheet, dtype=str)
		except ValueError as exc:
			# Handles files with Excel extension but CSV/plain-text content.
			if "format cannot be determined" in str(exc).lower():
				return pd.read_csv(path, dtype=str)
			raise

	raise ValueError(
		f"Unsupported file type '{path.suffix}' for file: {path}. "
		"Use CSV or Excel (.xlsx/.xls/.xlsm/.xlsb/.ods)."
	)


def compare_files(
	source_path: Path,
	target_path: Path,
	source_sheet: str | int,
	target_sheet: str | int,
	source_column: str | None,
	target_column: str | None,
) -> CompareResult:
	source_df = read_tabular_file(source_path, source_sheet)
	target_df = read_tabular_file(target_path, target_sheet)

	resolved_source_column = resolve_column(source_df, source_column)
	resolved_target_column = resolve_column(target_df, target_column)
	resolved_source_campus_column = resolve_column_with_candidates(
		source_df,
		preferred=None,
		candidates=DEFAULT_CAMPUS_CANDIDATES,
		error_hint="Could not auto-detect source campus column",
	)
	resolved_target_campus_column = resolve_column_with_candidates(
		target_df,
		preferred=None,
		candidates=DEFAULT_CAMPUS_CANDIDATES,
		error_hint="Could not auto-detect target campus column",
	)
	resolved_source_program_column = resolve_column_with_candidates(
		source_df,
		preferred=None,
		candidates=DEFAULT_SOURCE_PROGRAM_CANDIDATES,
		error_hint="Could not auto-detect source program_name column",
	)
	resolved_target_program_column = resolve_column_with_candidates(
		target_df,
		preferred=None,
		candidates=DEFAULT_TARGET_PROGRAM_CANDIDATES,
		error_hint="Could not auto-detect target program column",
	)
	resolved_source_college_column = resolve_column_with_candidates(
		source_df,
		preferred=None,
		candidates=DEFAULT_SOURCE_COLLEGE_CANDIDATES,
		error_hint="Could not auto-detect source college column",
	)
	resolved_target_college_columns = resolve_columns_with_candidates(
		target_df,
		candidates=DEFAULT_TARGET_COLLEGE_CANDIDATES,
	)
	if not resolved_target_college_columns:
		raise ValueError(
			"Could not auto-detect target college column. "
			"Expected one of: undergraduate_college, graduate_college, college. "
			f"Available columns: {list(target_df.columns)}"
		)
	source_duplicate_rows = get_duplicate_rows(source_df, resolved_source_column)
	target_duplicate_rows = get_duplicate_rows(target_df, resolved_target_column)

	source_names = extract_normalized_names(source_df, resolved_source_column)
	target_names = extract_normalized_names(target_df, resolved_target_column)

	source_set = set(source_names)
	target_set = set(target_names)

	found_set = sorted(target_set.intersection(source_set))
	not_found_set = sorted(target_set.difference(source_set))
	source_centered_rows = build_source_centered_found_not_found_df(
		source_names=source_names,
		target_names=target_names,
	)
	target_centered_rows = build_target_centered_found_not_found_df(
		target_names=target_names,
		source_names=source_names,
	)
	data_validation_rows = build_data_validation_rows(
		source_df=source_df,
		target_df=target_df,
		source_name_col=resolved_source_column,
		target_name_col=resolved_target_column,
		source_campus_col=resolved_source_campus_column,
		target_campus_col=resolved_target_campus_column,
		source_program_col=resolved_source_program_column,
		target_program_col=resolved_target_program_column,
		source_college_col=resolved_source_college_column,
		target_college_cols=resolved_target_college_columns,
		found_names=found_set,
	)

	return CompareResult(
		source_total_rows=len(source_df),
		target_total_rows=len(target_df),
		source_unique_names=len(source_set),
		target_unique_names=len(target_set),
		found_count=len(found_set),
		not_found_count=len(not_found_set),
		source_duplicate_names=count_duplicates(source_names),
		target_duplicate_names=count_duplicates(target_names),
		source_duplicate_rows=source_duplicate_rows,
		target_duplicate_rows=target_duplicate_rows,
		data_validation_rows=data_validation_rows,
		source_centered_rows=source_centered_rows,
		target_centered_rows=target_centered_rows,
		found_names=found_set,
		not_found_names=not_found_set,
	)


def build_default_output_path() -> Path:
	timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
	return OUTPUT_DIR / f"comparison_result_{timestamp}.xlsx"


def apply_data_validation_colors(data_validation_sheet, data_validation_df: pd.DataFrame) -> None:
	"""Apply color coding to validation columns in the Data Validation sheet.

	MATCH    -> Olive Green, Accent 3, Lighter 60%
	MISMATCH -> Red, Accent 2, Lighter 40%
	"""
	if data_validation_df.empty:
		return

	match_fill = PatternFill(fill_type="solid", fgColor=Color(theme=6, tint=0.6))
	mismatch_fill = PatternFill(fill_type="solid", fgColor=Color(theme=5, tint=0.4))

	validation_columns = [
		index + 1
		for index, col_name in enumerate(data_validation_df.columns)
		if str(col_name).strip().lower().endswith("_validation")
	]

	for excel_row in range(2, len(data_validation_df) + 2):
		for excel_col in validation_columns:
			cell = data_validation_sheet.cell(row=excel_row, column=excel_col)
			cell_value = str(cell.value).strip().upper() if cell.value is not None else ""
			if cell_value == "MATCH":
				cell.fill = match_fill
			elif cell_value == "MISMATCH":
				cell.fill = mismatch_fill


def build_combined_duplicates_df(result: CompareResult) -> pd.DataFrame:
	source_duplicates_df = result.source_duplicate_rows.copy()
	target_duplicates_df = result.target_duplicate_rows.copy()

	source_duplicates_df.insert(0, "file_type", "SOURCE")
	target_duplicates_df.insert(0, "file_type", "TARGET")

	combined_duplicates_df = pd.concat(
		[source_duplicates_df, target_duplicates_df],
		ignore_index=True,
		sort=False,
	)

	if combined_duplicates_df.empty:
		return pd.DataFrame(
			columns=[
				"file_type",
				"file_row_number",
				"normalized_fullname",
				"duplicate_count",
			]
		)

	if "file_row_number" in combined_duplicates_df.columns:
		combined_duplicates_df = combined_duplicates_df.sort_values(
			by=["file_type", "normalized_fullname", "file_row_number"],
			ascending=[True, True, True],
		)

	return combined_duplicates_df


def build_source_centered_found_not_found_df(
	source_names: list[str],
	target_names: list[str],
) -> pd.DataFrame:
	source_counter = Counter(source_names)
	target_counter = Counter(target_names)

	rows: list[dict[str, object]] = []
	for source_name in sorted(source_counter.keys()):
		target_occurrences = target_counter.get(source_name, 0)
		rows.append(
			{
				"source_fullname": source_name,
				"source_to_target_status": "FOUND" if target_occurrences > 0 else "NOT_FOUND",
			}
		)

	return pd.DataFrame(rows)


def build_target_centered_found_not_found_df(
	target_names: list[str],
	source_names: list[str],
) -> pd.DataFrame:
	target_counter = Counter(target_names)
	source_counter = Counter(source_names)

	rows: list[dict[str, object]] = []
	for target_name in sorted(target_counter.keys()):
		source_occurrences = source_counter.get(target_name, 0)
		rows.append(
			{
				"target_fullname": target_name,
				"target_to_source_status": "FOUND" if source_occurrences > 0 else "NOT_FOUND",
			}
		)

	return pd.DataFrame(rows)


def apply_found_not_found_colors(
	sheet,
	df: pd.DataFrame,
	status_column_name: str | None = None,
	default_status: str | None = None,
) -> None:
	if df.empty:
		return

	found_fill = PatternFill(fill_type="solid", fgColor=Color(theme=6, tint=0.6))
	not_found_fill = PatternFill(fill_type="solid", fgColor=Color(theme=5, tint=0.4))

	if status_column_name:
		status_col_idx = next(
			(i + 1 for i, col in enumerate(df.columns) if str(col).strip().lower() == status_column_name.strip().lower()),
			None,
		)
		if status_col_idx is None:
			return

		for row_idx in range(2, len(df) + 2):
			cell = sheet.cell(row=row_idx, column=status_col_idx)
			status = str(cell.value).strip().upper() if cell.value is not None else ""
			if status == "FOUND":
				cell.fill = found_fill
			elif status == "NOT_FOUND":
				cell.fill = not_found_fill
		return

	if default_status:
		first_col_idx = 1
		for row_idx in range(2, len(df) + 2):
			cell = sheet.cell(row=row_idx, column=first_col_idx)
			if default_status.upper() == "FOUND":
				cell.fill = found_fill
			elif default_status.upper() == "NOT_FOUND":
				cell.fill = not_found_fill


def autofit_all_worksheets(workbook) -> None:
	for sheet in workbook.worksheets:
		# Summary is a dashboard with intentional fixed sizing; don't auto-fit it.
		if sheet.title.strip().lower() == "summary":
			continue
		for column_cells in sheet.columns:
			max_length = 0
			column_index = column_cells[0].column
			for cell in column_cells:
				cell_value = "" if cell.value is None else str(cell.value)
				if len(cell_value) > max_length:
					max_length = len(cell_value)

			adjusted_width = min(max(max_length + 2, 12), 80)
			sheet.column_dimensions[get_column_letter(column_index)].width = adjusted_width


def freeze_top_row_all_worksheets(workbook) -> None:
	for sheet in workbook.worksheets:
		# Keep Summary more like a dashboard: freeze after the header + cards area.
		if sheet.title.strip().lower() == "summary":
			sheet.freeze_panes = "A8"
		else:
			sheet.freeze_panes = "A2"


def build_validation_summary_df(data_validation_df: pd.DataFrame) -> pd.DataFrame:
	"""Build a compact validation summary suitable for a dashboard-like Summary tab.

	We treat MATCH as good; MISMATCH as bad; MISSING/REVIEW as needs attention.
	"""
	areas = [
		("Campus validation", "campus_validation"),
		("Program validation", "program_validation"),
		("College validation", "college_validation"),
		("Overall validation", "overall_validation"),
	]

	rows: list[dict[str, object]] = []
	for label, col in areas:
		match = mismatch = attention = total = 0
		if not data_validation_df.empty and col in data_validation_df.columns:
			counts = (
				data_validation_df[col]
				.fillna("")
				.astype(str)
				.str.strip()
				.str.upper()
				.value_counts()
			)
			match = int(counts.get("MATCH", 0))
			mismatch = int(counts.get("MISMATCH", 0))
			attention = int(counts.get("MISSING", 0) + counts.get("REVIEW", 0))
			total = match + mismatch + attention

		match_rate = (match / total) if total else 0
		rows.append(
			{
				"validation_area": label,
				"match": match,
				"mismatch": mismatch,
				"attention": attention,
				"total": total,
				"match_rate": match_rate,
			}
		)

	return pd.DataFrame(rows)


def _style_cell_range(ws, cell_range: str, *, fill=None, font=None, alignment=None, border=None) -> None:
	for row in ws[cell_range]:
		for cell in row:
			if fill is not None:
				cell.fill = fill
			if font is not None:
				cell.font = font
			if alignment is not None:
				cell.alignment = alignment
			if border is not None:
				cell.border = border


def _write_card(
	ws,
	top_row: int,
	left_col: int,
	title: str,
	value: object,
	*,
	width_cols: int = 2,
	height_rows: int = 3,
	accent_rgb: str = "2563EB",
) -> None:
	"""Write a KPI-like 'card' using merged cells."""
	start_col = left_col
	end_col = left_col + width_cols - 1
	start_row = top_row
	end_row = top_row + height_rows - 1

	ws.merge_cells(
		start_row=start_row,
		start_column=start_col,
		end_row=end_row,
		end_column=end_col,
	)

	cell = ws.cell(row=start_row, column=start_col)
	cell.value = f"{title}\n{value}"
	cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
	cell.font = Font(name="Calibri", size=14, bold=True, color="111827")
	cell.fill = PatternFill(fill_type="solid", fgColor="F8FAFC")

	thin = Side(style="thin", color="E5E7EB")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for r in range(start_row, end_row + 1):
		for c in range(start_col, end_col + 1):
			ws.cell(row=r, column=c).border = border

	# Accent line (top border) by coloring the first row's top border
	accent_side = Side(style="medium", color=accent_rgb)
	for c in range(start_col, end_col + 1):
		ws.cell(row=start_row, column=c).border = Border(
			left=thin,
			right=thin,
			top=accent_side,
			bottom=thin,
		)


def build_modern_summary_sheet(
	summary_sheet,
	*,
	result: CompareResult,
	source_label: str,
	target_label: str,
) -> None:
	"""Create a clean, modern Summary dashboard with charts."""
	ws = summary_sheet
	ws.sheet_view.showGridLines = False
	ws.page_setup.fitToWidth = 1
	ws.page_setup.fitToHeight = 0

	# Column widths tuned for a dashboard layout.
	for col, width in {
		"A": 22,
		"B": 14,
		"C": 22,
		"D": 14,
		"E": 22,
		"F": 14,
		"G": 22,
		"H": 14,
	}.items():
		ws.column_dimensions[col].width = width

	# Title
	ws.merge_cells("A1:H1")
	ws["A1"].value = "Validation Summary"
	ws["A1"].font = Font(name="Calibri", size=20, bold=True, color="111827")
	ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
	ws.row_dimensions[1].height = 30

	ws.merge_cells("A2:H2")
	ws["A2"].value = (
		f"Source: {source_label}    |    Target: {target_label}    |    Generated: "
		f"{datetime.now().strftime('%Y-%m-%d %H:%M')}"
	)
	ws["A2"].font = Font(name="Calibri", size=11, color="6B7280")
	ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
	ws.row_dimensions[2].height = 18

	# KPI cards
	target_to_source_found_count = int(
		(result.target_centered_rows["target_to_source_status"] == "FOUND").sum()
	) if not result.target_centered_rows.empty else 0
	target_to_source_not_found_count = int(
		(result.target_centered_rows["target_to_source_status"] == "NOT_FOUND").sum()
	) if not result.target_centered_rows.empty else 0

	_write_card(ws, 4, 1, "Source rows", result.source_total_rows, accent_rgb="2563EB")
	_write_card(ws, 4, 3, "Target rows", result.target_total_rows, accent_rgb="7C3AED")
	_write_card(ws, 4, 5, "Names FOUND", target_to_source_found_count, accent_rgb="16A34A")
	_write_card(ws, 4, 7, "Names NOT FOUND", target_to_source_not_found_count, accent_rgb="DC2626")
	for r in range(4, 7):
		ws.row_dimensions[r].height = 22

	# Section header
	ws.merge_cells("A8:H8")
	ws["A8"].value = "Data Validation Summary"
	ws["A8"].font = Font(name="Calibri", size=13, bold=True, color="111827")
	ws["A8"].alignment = Alignment(horizontal="left", vertical="center")
	ws.row_dimensions[8].height = 22

	# Build the validation summary table
	validation_summary_df = build_validation_summary_df(result.data_validation_rows)

	start_row = 9
	headers = ["Validation area", "MATCH", "MISMATCH", "MISSING/REVIEW", "TOTAL", "MATCH RATE"]
	for idx, header in enumerate(headers, start=1):
		cell = ws.cell(row=start_row, column=idx)
		cell.value = header
		cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
		cell.fill = PatternFill(fill_type="solid", fgColor="111827")
		cell.alignment = Alignment(horizontal="center", vertical="center")
	ws.row_dimensions[start_row].height = 20

	thin = Side(style="thin", color="E5E7EB")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)

	for row_offset, record in enumerate(validation_summary_df.to_dict(orient="records"), start=1):
		r = start_row + row_offset
		ws.cell(row=r, column=1, value=record["validation_area"]).alignment = Alignment(
			horizontal="left", vertical="center"
		)
		ws.cell(row=r, column=2, value=int(record["match"]))
		ws.cell(row=r, column=3, value=int(record["mismatch"]))
		ws.cell(row=r, column=4, value=int(record["attention"]))
		ws.cell(row=r, column=5, value=int(record["total"]))
		ws.cell(row=r, column=6, value=float(record["match_rate"]))
		ws.cell(row=r, column=6).number_format = "0%"
		ws.row_dimensions[r].height = 18

		# Subtle zebra striping for modern look
		if row_offset % 2 == 1:
			fill = PatternFill(fill_type="solid", fgColor="F9FAFB")
			_style_cell_range(ws, f"A{r}:F{r}", fill=fill)

		for c in range(1, 7):
			cell = ws.cell(row=r, column=c)
			cell.border = border
			if c in {2, 3, 4, 5, 6}:
				cell.alignment = Alignment(horizontal="center", vertical="center")
			else:
				cell.alignment = Alignment(horizontal="left", vertical="center")

	# Border header row
	for c in range(1, 7):
		ws.cell(row=start_row, column=c).border = border

	# Conditional color highlights (modern, non-neon)
	match_fill = PatternFill(fill_type="solid", fgColor="DCFCE7")
	mismatch_fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
	attention_fill = PatternFill(fill_type="solid", fgColor="FEF9C3")
	for r in range(start_row + 1, start_row + 1 + len(validation_summary_df)):
		ws.cell(row=r, column=2).fill = match_fill
		ws.cell(row=r, column=3).fill = mismatch_fill
		ws.cell(row=r, column=4).fill = attention_fill

	# Duplicates section
	dup_section_row = start_row + 1 + len(validation_summary_df) + 2  # after table + spacing
	ws.merge_cells(f"A{dup_section_row}:H{dup_section_row}")
	ws[f"A{dup_section_row}"].value = "Duplicates"
	ws[f"A{dup_section_row}"].font = Font(name="Calibri", size=13, bold=True, color="111827")
	ws[f"A{dup_section_row}"].alignment = Alignment(horizontal="left", vertical="center")
	ws.row_dimensions[dup_section_row].height = 22

	dup_header_row = dup_section_row + 1
	dup_headers = ["Metric", "Source", "Target"]
	for idx, header in enumerate(dup_headers, start=1):
		cell = ws.cell(row=dup_header_row, column=idx, value=header)
		cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
		cell.fill = PatternFill(fill_type="solid", fgColor="111827")
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = border
	ws.row_dimensions[dup_header_row].height = 20

	dup_rows = [
		("Duplicate names", int(result.source_duplicate_names), int(result.target_duplicate_names)),
		("Duplicate rows", int(len(result.source_duplicate_rows)), int(len(result.target_duplicate_rows))),
	]
	for i, (metric, src_val, tgt_val) in enumerate(dup_rows, start=1):
		r = dup_header_row + i
		ws.cell(row=r, column=1, value=metric).alignment = Alignment(horizontal="left", vertical="center")
		ws.cell(row=r, column=2, value=src_val).alignment = Alignment(horizontal="center", vertical="center")
		ws.cell(row=r, column=3, value=tgt_val).alignment = Alignment(horizontal="center", vertical="center")
		ws.row_dimensions[r].height = 18
		if i % 2 == 1:
			_style_cell_range(ws, f"A{r}:C{r}", fill=PatternFill(fill_type="solid", fgColor="F9FAFB"))
		for c in range(1, 4):
			ws.cell(row=r, column=c).border = border

	# Matching coverage section
	match_section_row = dup_header_row + len(dup_rows) + 3
	ws.merge_cells(f"A{match_section_row}:H{match_section_row}")
	ws[f"A{match_section_row}"].value = "Name matching coverage"
	ws[f"A{match_section_row}"].font = Font(name="Calibri", size=13, bold=True, color="111827")
	ws[f"A{match_section_row}"].alignment = Alignment(horizontal="left", vertical="center")
	ws.row_dimensions[match_section_row].height = 22

	match_header_row = match_section_row + 1
	match_headers = ["Direction", "FOUND", "NOT FOUND", "FOUND RATE"]
	for idx, header in enumerate(match_headers, start=1):
		cell = ws.cell(row=match_header_row, column=idx, value=header)
		cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
		cell.fill = PatternFill(fill_type="solid", fgColor="111827")
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = border
	ws.row_dimensions[match_header_row].height = 20

	source_to_target_found_count = int(
		(result.source_centered_rows["source_to_target_status"] == "FOUND").sum()
	) if not result.source_centered_rows.empty else 0
	source_to_target_not_found_count = int(
		(result.source_centered_rows["source_to_target_status"] == "NOT_FOUND").sum()
	) if not result.source_centered_rows.empty else 0

	def _rate(found: int, not_found: int) -> float:
		total = found + not_found
		return (found / total) if total else 0

	match_rows = [
		(
			"Target → Source",
			target_to_source_found_count,
			target_to_source_not_found_count,
			_rate(target_to_source_found_count, target_to_source_not_found_count),
		),
		(
			"Source → Target",
			source_to_target_found_count,
			source_to_target_not_found_count,
			_rate(source_to_target_found_count, source_to_target_not_found_count),
		),
	]
	for i, (direction, found, not_found, rate) in enumerate(match_rows, start=1):
		r = match_header_row + i
		ws.cell(row=r, column=1, value=direction).alignment = Alignment(horizontal="left", vertical="center")
		ws.cell(row=r, column=2, value=int(found)).alignment = Alignment(horizontal="center", vertical="center")
		ws.cell(row=r, column=3, value=int(not_found)).alignment = Alignment(horizontal="center", vertical="center")
		ws.cell(row=r, column=4, value=float(rate)).number_format = "0%"
		ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
		ws.row_dimensions[r].height = 18
		if i % 2 == 1:
			_style_cell_range(ws, f"A{r}:D{r}", fill=PatternFill(fill_type="solid", fgColor="F9FAFB"))
		for c in range(1, 5):
			ws.cell(row=r, column=c).border = border

	# Footer notes
	footer_row = match_header_row + len(match_rows) + 3
	ws.merge_cells(f"A{footer_row}:H{footer_row}")
	ws[f"A{footer_row}"].value = "Tip: Use 'Duplicates', 'Target_to_Source', 'Source_to_Target', and 'Data Validation' tabs to drill down."
	ws[f"A{footer_row}"].font = Font(name="Calibri", size=10, color="6B7280")
	ws[f"A{footer_row}"].alignment = Alignment(horizontal="left", vertical="center")
	ws.row_dimensions[footer_row].height = 18


def write_report_to_file(
	result: CompareResult,
	output_path: Path,
	*,
	source_label: str = "SOURCE",
	target_label: str = "TARGET",
) -> None:
	if output_path.suffix.lower() != ".xlsx":
		raise ValueError("Output file must use .xlsx extension.")

	target_to_source_df = result.target_centered_rows
	combined_duplicates_df = build_combined_duplicates_df(result)

	output_path.parent.mkdir(parents=True, exist_ok=True)
	with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
		target_to_source_df.to_excel(writer, sheet_name="Target_to_Source", index=False)
		result.source_centered_rows.to_excel(writer, sheet_name="Source_to_Target", index=False)
		combined_duplicates_df.to_excel(writer, sheet_name="Duplicates", index=False)
		result.data_validation_rows.to_excel(writer, sheet_name="Data Validation", index=False)
		# Rebuild Summary from scratch (clean/modern) so old user-added charts/analytics
		# are not carried over.
		wb = writer.book
		if "Summary" in wb.sheetnames:
			wb.remove(wb["Summary"])
		summary_sheet = wb.create_sheet("Summary", 0)
		writer.sheets["Summary"] = summary_sheet

		source_to_target_sheet = writer.sheets["Source_to_Target"]
		target_to_source_sheet = writer.sheets["Target_to_Source"]
		data_validation_sheet = writer.sheets["Data Validation"]
		build_modern_summary_sheet(
			summary_sheet,
			result=result,
			source_label=source_label,
			target_label=target_label,
		)
		apply_found_not_found_colors(
			source_to_target_sheet,
			result.source_centered_rows,
			status_column_name="source_to_target_status",
		)
		apply_found_not_found_colors(
			target_to_source_sheet,
			target_to_source_df,
			status_column_name="target_to_source_status",
		)
		apply_data_validation_colors(data_validation_sheet, result.data_validation_rows)
		freeze_top_row_all_worksheets(writer.book)
		autofit_all_worksheets(writer.book)


def parse_sheet_arg(value: str | int) -> str | int:
	if isinstance(value, int):
		return value
	text = str(value).strip()
	if text.isdigit():
		return int(text)
	return text


def main() -> None:
	args = parse_args()

	ensure_managed_directories()

	source_path = resolve_input_path(args.source, SOURCE_DATA_DIR)
	target_path = resolve_input_path(args.target, TARGET_DATA_DIR)

	if not source_path.exists():
		raise FileNotFoundError(f"Source file not found: {source_path}")
	if not target_path.exists():
		raise FileNotFoundError(f"Target file not found: {target_path}")

	output_path = resolve_output_path(args.output)

	result = compare_files(
		source_path=source_path,
		target_path=target_path,
		source_sheet=parse_sheet_arg(args.sheet_source),
		target_sheet=parse_sheet_arg(args.sheet_target),
		source_column=args.source_column,
		target_column=args.target_column,
	)
	write_report_to_file(
		result,
		output_path,
		source_label=source_path.name,
		target_label=target_path.name,
	)
	print(f"Comparison report created: {output_path.resolve()}")


if __name__ == "__main__":
	main()
